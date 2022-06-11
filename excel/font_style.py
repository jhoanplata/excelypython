from openpyxl.styles import Font 
from openpyxl.styles.colors import Color
from openpyxl import Workbook

# Create workbook
wb = Workbook()

# Select sheet
ws = wb['Sheet']

# Set font
ws['A1'].font = Font(color = "FF0000")

# Set data
ws['A1'] = "Hello frien"
    
# Save file    
wb.save("styles.xlsx")