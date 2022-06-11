from os import system
from re import X
system("clear")

from openpyxl import load_workbook

wb = load_workbook('sample.xlsx')
print(wb.sheetnames)


#se debe instalar openpyxl para poder iniciar