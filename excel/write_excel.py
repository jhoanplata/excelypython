from os import system
from re import X
system("clear")

from openpyxl import Workbook 

wb = Workbook()
ws = wb.active

ws["A2"] = "Hello"
ws["B12"] = "World"
ws["C4"] = "from"
ws["D7"] = "Python"

wb.save("sample2.xlsx")
